import streamlit as st
import dropbox
import openpyxl
import os
import re
import asyncio
import aiohttp
import json
import random
import pandas as pd
from collections import defaultdict
from io import BytesIO

# --- Page Configuration ---
st.set_page_config(
    page_title="Dropbox Batch Link Exporter Pro",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.title("📦 Dropbox Batch Link Exporter (Pro)")
st.markdown("""
**Reliability & Error Handling Update:** 
This tool now creates a separate "Error Report" Excel file and keeps your results visible even after downloading files.
""")

# --- Initialize Session State (Persistence Layer) ---
keys_to_init = [
    'connection_success', 'dbx_token', 'current_path', 'multi_select_list',
    'processing_done', 'success_data', 'failure_data', 
    'excel_success_buffer', 'excel_error_buffer',
    'ordered_file_rows',
    'custom_filename_input' # NEW: Store the user's preferred filename
]

for key in keys_to_init:
    if key not in st.session_state:
        if key in ['multi_select_list', 'success_data', 'failure_data', 'ordered_file_rows']:
            st.session_state[key] = []
        elif 'buffer' in key: st.session_state[key] = None
        elif 'done' in key or 'success' in key: st.session_state[key] = False
        elif 'input' in key: st.session_state[key] = "Dropbox_Links_Export" # Default name
        else: st.session_state[key] = ""

if st.session_state.current_path == "":
    st.session_state.current_path = "/"


# --- Sidebar ---
with st.sidebar:
    st.header("1. Configuration")
    DROPBOX_ACCESS_TOKEN = st.text_input(
        "Dropbox Access Token",
        type="password",
        value=st.session_state.dbx_token,
        help="Enter your Dropbox API access token."
    )
    st.session_state.dbx_token = DROPBOX_ACCESS_TOKEN

    st.header("2. Tuning Parameters")
    MAX_CONCURRENT_TASKS = st.slider("Max Concurrent Tasks (Global Limit)", 1, 30, 15, help="This is the maximum limit. If you process fewer files, the app will automatically lower this number.") 
    MAX_RETRIES = st.slider("Max Retries per File", 1, 10, 5)

    st.divider()
    if st.button("🔄 Reset App State"):
        for key in st.session_state.keys():
            del st.session_state[key]
        st.rerun()
    st.divider()
    with st.expander("📖 View Documentation"):
        st.subheader("One-Time Setup: Getting your Dropbox Token")
        st.markdown("""
        To use this tool, you need a special access token from Dropbox. Follow these steps exactly:
        1.  **Go to the Dropbox App Console**: Log in to your Dropbox account and navigate to [https://www.dropbox.com/developers/apps](https://www.dropbox.com/developers/apps).
        2.  **Create a New App**:
            - Click the **"Create app"** button.
            - Choose **"Scoped access"**.
            - Select **"Full Dropbox"** as the type of access.
            - Give your app a unique name (e.g., "My Streamlit Link Exporter").
        3.  **Configure Permissions**:
            - After creating the app, you will be taken to its settings page. Go to the **"Permissions"** tab.
            - You **must** check the boxes for the following permissions:
                - `files.metadata.read` (To see files and folders)
                - `sharing.write` (To create shareable links)
                - `sharing.read` (To read existing shareable links)
            - Click **"Submit"** at the bottom to save the changes.
        4.  **Generate Access Token**:
            - Go back to the **"Settings"** tab.
            - Find the "Generated access token" section.
            - By default, the token expires in 4 hours. For longer sessions, you can change this by selecting **"No expiration"** from the dropdown.
            - Click the **"Generate"** button.
        5.  **Copy and Use**: Copy the long string of characters. This is your Access Token. Paste it into the "Dropbox Access Token" field in this app's sidebar.
        """)

        st.subheader("How to Use This Tool")
        st.markdown("""
        1.  **Enter Token**: Paste your generated Dropbox Access Token into the "Configuration" section of the sidebar.
        2.  **Connect**: Click the **"Connect"** button on the main page. A green success message will appear if the token is valid.
        3.  **Navigate & Select**:
            - A folder browser will appear. Click **"Open"** to enter a subfolder.
            - Use the **"Up Level"** button to navigate back.
            - Use the **checkboxes** next to the folder names to select one or more folders for processing.
            - Your selected folders will appear in the "Selected Folders" expander.
        4.  **Process Files**:
            - Once selected, give your output Excel file a custom name (optional).
            - Click the **"Start Processing"** button.
        5.  **Download Results**:
            - Wait for the progress bar to complete.
            - **Success File**: Contains all working links.
            - **Error Report**: (If applicable) Contains details on any files that failed.
        """)
# --- CORE LOGIC ---

def list_dropbox_folders(token, path):
    api_path = "" if path == "/" else path
    try:
        dbx = dropbox.Dropbox(token)
        result = dbx.files_list_folder(path=api_path)
        folders = [entry for entry in result.entries if isinstance(entry, dropbox.files.FolderMetadata)]
        while result.has_more:
            result = dbx.files_list_folder_continue(result.cursor)
            folders.extend([entry for entry in result.entries if isinstance(entry, dropbox.files.FolderMetadata)])
        return sorted(folders, key=lambda x: x.name.lower())
    except dropbox.exceptions.AuthError:
        st.error("Authentication Error: Token is invalid.")
        st.session_state.connection_success = False
        return None
    except Exception as e:
        st.error(f"Error listing folders: {e}")
        return None

def get_all_file_metadata(folder_path, token, status_placeholder):
    api_path = "" if folder_path == "/" else folder_path
    try:
        dbx = dropbox.Dropbox(token)
        status_placeholder.info(f"Scanning '{folder_path}'...")
        all_files, result = [], dbx.files_list_folder(api_path, recursive=True)
        all_files.extend(entry for entry in result.entries if isinstance(entry, dropbox.files.FileMetadata))
        while result.has_more:
            result = dbx.files_list_folder_continue(result.cursor)
            all_files.extend(entry for entry in result.entries if isinstance(entry, dropbox.files.FileMetadata))
        return [f for f in all_files if not f.name.lower().endswith(('.xlsx', '.docx', '.pdf', '.zip', '.db', '.ds_store'))]
    except Exception as e:
        st.warning(f"Could not scan {folder_path}: {e}")
        return []


# --- NEW: derive the "child SKU" (Parent-ColorCode) directly from the filename ---
# Expected filename pattern: <ParentName>-<ColorCode>_<sequence>.<ext>
# e.g. "Sofa123-Red_1.jpg", "Sofa123-Red_2.jpg" -> child SKU "Sofa123-Red"
# This is what makes each parent+color combination unique, instead of the
# color folder name alone (which repeats across every parent SKU).
CHILD_SKU_PATTERN = re.compile(r'_\d+$')

def derive_child_sku(image_name: str) -> str:
    name_no_ext = os.path.splitext(image_name)[0]
    child_sku = CHILD_SKU_PATTERN.sub('', name_no_ext).strip()
    return child_sku if child_sku else name_no_ext


async def worker(name, queue, session, token, semaphore):
    async with semaphore: 
        while not queue.empty():
            entry = await queue.get()
            full_path = os.path.dirname(entry.path_display)
            subfolder = os.path.basename(full_path)
            image_name = entry.name
            
            headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
            create_url = 'https://api.dropboxapi.com/2/sharing/create_shared_link_with_settings'
            list_url = 'https://api.dropboxapi.com/2/sharing/list_shared_links'
            
            success = False
            error_msg = "Unknown"

            for attempt in range(MAX_RETRIES):
                try:
                    create_payload = json.dumps({'path': entry.path_lower})
                    async with session.post(create_url, headers=headers, data=create_payload) as response:
                        if response.status == 200:
                            data = await response.json()
                            st.session_state.success_data.append([full_path, subfolder, image_name, data.get('url'), entry.path_lower])
                            success = True
                            break
                        elif response.status == 409:
                            list_payload = json.dumps({'path': entry.path_lower})
                            async with session.post(list_url, headers=headers, data=list_payload) as list_resp:
                                if list_resp.status == 200:
                                    list_data = await list_resp.json()
                                    if list_data.get('links'):
                                        st.session_state.success_data.append([full_path, subfolder, image_name, list_data['links'][0]['url'], entry.path_lower])
                                        success = True
                                        break
                                    else:
                                        error_msg = "Conflict (409) but no links returned."
                        elif response.status == 429:
                            retry_after = int(response.headers.get("Retry-After", 1))
                            wait_time = retry_after + (random.uniform(0.5, 2.0))
                            await asyncio.sleep(wait_time)
                            continue 
                        elif response.status >= 500:
                             error_msg = f"Server Error {response.status}"
                        else:
                            try:
                                txt = await response.text()
                                error_msg = f"HTTP {response.status}: {txt}"
                            except:
                                error_msg = f"HTTP {response.status}"
                except Exception as e:
                    error_msg = f"Exception: {str(e)}"
                
                await asyncio.sleep((1.5 ** attempt) + random.uniform(0.5, 1.5))
            
            if not success:
                st.session_state.failure_data.append([full_path, subfolder, image_name, error_msg, entry.path_lower])
            queue.task_done()

async def run_file_queue(file_list, token, progress_label, concurrency):
    if not file_list: return

    queue = asyncio.Queue()
    for f in file_list: queue.put_nowait(f)
    
    total_files = len(file_list)
    
    effective_concurrency = min(total_files, concurrency)
    start_success_count = len(st.session_state.success_data)
    start_failure_count = len(st.session_state.failure_data)

    st.info(f"{progress_label}: using {effective_concurrency} concurrent workers for {total_files} files.")
    progress_bar = st.progress(0, text=f"Initializing...")
    
    semaphore = asyncio.Semaphore(effective_concurrency)
    timeout = aiohttp.ClientTimeout(total=600, connect=60) 
    connector = aiohttp.TCPConnector(limit=effective_concurrency, force_close=True) 

    async with aiohttp.ClientSession(timeout=timeout, connector=connector) as session:
        tasks = [asyncio.create_task(worker(f'w-{i}', queue, session, token, semaphore)) for i in range(effective_concurrency)]
        
        while not queue.empty():
            processed = total_files - queue.qsize()
            pass_success = len(st.session_state.success_data) - start_success_count
            pass_failure = len(st.session_state.failure_data) - start_failure_count
            progress_bar.progress(
                processed / total_files, 
                text=f"{progress_label}: {processed}/{total_files} | Success: {pass_success} | Fail: {pass_failure}"
            )
            await asyncio.sleep(0.5)
        
        await queue.join()
        for task in tasks: task.cancel()
        
    progress_bar.progress(1.0, text=f"{progress_label} complete.")

async def process_files_logic(all_files, token):
    if not all_files: return

    st.session_state.success_data = []
    st.session_state.failure_data = []
    ordered_source_files = sorted(
        all_files,
        key=lambda f: (
            os.path.basename(os.path.dirname(f.path_display)).lower(),
            f.name.lower()
        )
    )
    st.session_state.ordered_file_rows = [
        [
            os.path.dirname(f.path_display),
            os.path.basename(os.path.dirname(f.path_display)),
            f.name,
            f.path_lower
        ]
        for f in ordered_source_files
    ]
    
    # --- DYNAMIC CONCURRENCY LOGIC ---
    # If we have 4 files, we use 4 workers.
    # If we have 100 files, we use MAX_CONCURRENT_TASKS (e.g., 15).
    await run_file_queue(all_files, token, "Pass 1 - processing all files", MAX_CONCURRENT_TASKS)

    if st.session_state.failure_data:
        failed_paths = {row[4] for row in st.session_state.failure_data}
        failed_files = [f for f in all_files if f.path_lower in failed_paths]
        st.warning(f"Retrying {len(failed_files)} failed file(s) one more time before generating Excel.")
        st.session_state.failure_data = []
        retry_concurrency = 1
        await run_file_queue(failed_files, token, "Pass 2 - retrying failed files", retry_concurrency)

    st.info("Network operations complete! Generating Excel...")

def generate_excel_buffers():
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Master List"
    # NEW: Child SKU column added so it's visible/auditable in the Master List too
    ws_main.append(["Full Path", "Subfolder (Color)", "Image Name", "Child SKU", "Direct Link"])

    def clean_link(link):
        if not link: return ""
        if 'dl=0' in link: return link.replace('dl=0', 'dl=1')
        return link + '&dl=1' if '?' in link else link + '?dl=1'

    success_by_path = {}
    for row in st.session_state.success_data:
        path_key = row[4] if len(row) > 4 else f"{row[0]}/{row[2]}".lower()
        success_by_path[path_key] = row

    ordered_rows = st.session_state.ordered_file_rows
    if not ordered_rows:
        ordered_rows = [
            [row[0], row[1], row[2], row[4] if len(row) > 4 else f"{row[0]}/{row[2]}".lower()]
            for row in st.session_state.success_data
        ]

    # NEW: build child-sku-keyed groups instead of subfolder(color)-keyed groups
    child_sku_map = defaultdict(list)

    for full_path, subfolder, image_name, path_key in ordered_rows:
        link = clean_link(success_by_path[path_key][3]) if path_key in success_by_path else ""
        child_sku = derive_child_sku(image_name)
        ws_main.append([full_path, subfolder, image_name, child_sku, link])
        child_sku_map[child_sku].append(link)

    # Sort images within each child SKU by their natural filename order (they were
    # already appended in the ordered_rows sequence, which is sorted by subfolder+name,
    # so this preserves _1, _2, _3... order as long as filenames sort naturally).

    ws_horiz = wb.create_sheet("Horizontal View")
    if child_sku_map:
        max_imgs = max(len(v) for v in child_sku_map.values())
        ws_horiz.append(["Child SKU"] + [f"Image {i+1}" for i in range(max_imgs)])
        for child_sku in sorted(child_sku_map.keys()):
            ws_horiz.append([child_sku] + child_sku_map[child_sku])

    success_buffer = BytesIO()
    wb.save(success_buffer)
    success_buffer.seek(0)
    st.session_state.excel_success_buffer = success_buffer

    if st.session_state.failure_data:
        wb_err = openpyxl.Workbook()
        ws_err = wb_err.active
        ws_err.title = "Failed Files"
        ws_err.append(["Folder Path", "Subfolder (Color)", "File Name", "Error Reason"])
        
        st.session_state.failure_data.sort(key=lambda x: x[1])
        for row in st.session_state.failure_data:
            ws_err.append(row[:4])
            
        error_buffer = BytesIO()
        wb_err.save(error_buffer)
        error_buffer.seek(0)
        st.session_state.excel_error_buffer = error_buffer
    else:
        st.session_state.excel_error_buffer = None


# --- UI NAVIGATION ---

if not st.session_state.connection_success:
    st.subheader("Step 1: Connect to Dropbox")
    if st.button("Connect", use_container_width=True):
        if not st.session_state.dbx_token:
            st.warning("Please enter Token in sidebar.")
        else:
            with st.spinner("Connecting..."):
                if list_dropbox_folders(st.session_state.dbx_token, "/") is not None:
                    st.session_state.connection_success = True; st.rerun()

else:
    st.subheader("Step 2: Navigate and Select Folders")
    
    def go_up(): st.session_state.current_path = os.path.dirname(st.session_state.current_path)
    def go_in(p): st.session_state.current_path = p
    def toggle(p):
        if p in st.session_state.multi_select_list: st.session_state.multi_select_list.remove(p)
        else: st.session_state.multi_select_list.append(p)

    col_nav1, col_nav2 = st.columns([1,5])
    col_nav1.button("⏫ Up Level", on_click=go_up, disabled=(st.session_state.current_path=="/"))
    col_nav2.info(f"Path: `{st.session_state.current_path}`")
    
    subfolders = list_dropbox_folders(st.session_state.dbx_token, st.session_state.current_path)
    if subfolders is not None:
        if subfolders:
            for f in subfolders:
                c1, c2 = st.columns([4,1])
                sel = f.path_display in st.session_state.multi_select_list
                c1.checkbox(f"📁 {f.name}", value=sel, key=f"chk_{f.id}", on_change=toggle, args=(f.path_display,))
                c2.button("Open", key=f"btn_{f.id}", on_click=go_in, args=(f.path_display,))
        else:
            st.write("No folders found here.")

    with st.expander(f"📚 Selected Folders ({len(st.session_state.multi_select_list)})", expanded=False):
        st.write(st.session_state.multi_select_list)
        if st.button("Clear Selection"): st.session_state.multi_select_list = []; st.rerun()


    st.divider()
    st.subheader("Step 3: Process")
    
    # --- CUSTOM FILENAME INPUT ---
    st.session_state.custom_filename_input = st.text_input(
        "📝 Name your output Excel file:", 
        value=st.session_state.custom_filename_input,
        help="Do not include .xlsx, just the name."
    )

    if not st.session_state.processing_done:
        if st.button("🚀 Start Processing", type="primary", disabled=not st.session_state.multi_select_list):
            
            all_files = []
            processed_paths = set()
            status_area = st.empty()
            
            for folder_path in st.session_state.multi_select_list:
                files = get_all_file_metadata(folder_path, st.session_state.dbx_token, status_area)
                for f in files:
                    if f.path_lower not in processed_paths:
                        all_files.append(f)
                        processed_paths.add(f.path_lower)
            
            if not all_files:
                st.error("No files found in selected folders.")
            else:
                asyncio.run(process_files_logic(all_files, st.session_state.dbx_token))
                generate_excel_buffers()
                st.session_state.processing_done = True
                st.rerun()
    
    if st.session_state.processing_done:
        st.success("Processing Complete!")
        
        unique_success_paths = {
            row[4] if len(row) > 4 else f"{row[0]}/{row[2]}".lower()
            for row in st.session_state.success_data
        }
        succ_count = len(unique_success_paths)
        fail_count = len(st.session_state.failure_data)
        total_count = len(st.session_state.ordered_file_rows) or (succ_count + fail_count)
        
        m1, m2, m3 = st.columns(3)
        m1.metric("✅ Success", succ_count)
        m2.metric("❌ Failed", fail_count)
        m3.metric("Total", total_count)
        
        # Format filename logic
        base_name = st.session_state.custom_filename_input.strip()
        if not base_name: base_name = "Dropbox_Export"
        success_filename = f"{base_name}.xlsx"
        error_filename = f"{base_name}_ERRORS.xlsx"

        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                label=f"📥 Download {success_filename}",
                data=st.session_state.excel_success_buffer,
                file_name=success_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with d2:
            if st.session_state.excel_error_buffer:
                st.download_button(
                    label="⚠️ Download ERROR Report",
                    data=st.session_state.excel_error_buffer,
                    file_name=error_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.button("No Errors to Download", disabled=True, use_container_width=True)

        if fail_count > 0:
            st.divider()
            st.error("⚠️ Attention Needed: Missing Files Analysis")
            
            df_fail = pd.DataFrame(
                [row[:4] for row in st.session_state.failure_data],
                columns=["Path", "SKU (Subfolder)", "File", "Reason"]
            )
            sku_counts = df_fail['SKU (Subfolder)'].value_counts().reset_index()
            sku_counts.columns = ['SKU Name', 'Missing Image Count']
            
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown("#### 🚨 Problematic SKUs")
                st.dataframe(sku_counts, hide_index=True, use_container_width=True)
            
            with c2:
                st.markdown("#### 🔍 Detailed Error Log")
                sku_options = ["All"] + sorted(list(sku_counts['SKU Name']))
                selected_sku = st.selectbox("Filter Errors by SKU:", sku_options)
                
                if selected_sku != "All":
                    display_df = df_fail[df_fail['SKU (Subfolder)'] == selected_sku]
                else:
                    display_df = df_fail
                
                st.dataframe(display_df[['File', 'Reason', 'SKU (Subfolder)']], hide_index=True, use_container_width=True)
                
        st.divider()
        if st.button("Start New Batch (Clear All Results)"):
            st.session_state.processing_done = False
            st.session_state.success_data = []
            st.session_state.failure_data = []
            st.session_state.ordered_file_rows = []
            st.session_state.excel_success_buffer = None
            st.session_state.excel_error_buffer = None
            st.rerun()

st.markdown("""
<style>
    footer {visibility: hidden;}
    .footer-text {text-align: center; color: #808080; padding: 20px;}
</style>
<div class="footer-text">
    <p>© 2025 Virtual Ops. Developed by Mohit Dhaker</p>
</div>
""", unsafe_allow_html=True)